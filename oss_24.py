import requests
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl as op
import os
from openpyxl.styles import Alignment, Font
from openpyxl.styles import PatternFill

now = datetime.now().strftime("%Y-%m-%d")
now2 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

urls = [
    "https://www.kma.go.kr/wid/queryDFSRSS.jsp?zone=3017066000",
    "https://www.kma.go.kr/wid/queryDFSRSS.jsp?zone=1168066000",
    "https://www.kma.go.kr/wid/queryDFSRSS.jsp?zone=3114056000",
    "https://www.kma.go.kr/wid/queryDFSRSS.jsp?zone=2644053000",
    "https://www.kma.go.kr/wid/queryDFSRSS.jsp?zone=2920054000",
    "https://www.kma.go.kr/wid/queryDFSRSS.jsp?zone=5111035000"
]

url2 = "https://news.sbs.co.kr/news/keywordList.do?keyword=%EB%82%A0%EC%94%A8&plink=TOPWORD&cooper=SBSNEWSKEYWORD"

headers = {
    'User-Agent': 'Mozilla/5.0',
    'Content-Type': 'text/html; charset=utf-8'
}
r = requests.get(url2, headers=headers)

soup = BeautifulSoup(r.text, 'lxml')
tags = soup.select('#container > div > div.w_news_list.type_issue2 > ul > li > a > p > strong')

a=[]
for tag in tags:
    title = tag.get_text()
    a.append(title)
filtered_a = [item for item in a if '[날씨]' in item]

def fetch_weather_data(url):
    req = requests.get(url, headers=headers)
    soup = BeautifulSoup(req.text, "xml")

    category = soup.find("category").string.split()[0]

    current_hour = datetime.now().hour
    closest_hour_data = None
    smallest_diff = float('inf')

    for item in soup.find_all("data"):
        hour = int(item.find("hour").string)
        diff = abs(current_hour - hour)
        if diff < smallest_diff:
            smallest_diff = diff
            closest_hour_data = item

    if closest_hour_data:
        hour = closest_hour_data.find("hour").string
        temp = closest_hour_data.find("temp").string
        reh = closest_hour_data.find("reh").string
        pop = closest_hour_data.find("pop").string if closest_hour_data.find("pop") else "0"
        wfKor = closest_hour_data.find("wfKor").string
        r06 = closest_hour_data.find("r06").string
        
        weather_info = {
            "category": category,
            "hour": hour,
            "temp": temp,
            "reh": reh,
            "pop": pop,
            "wfKor": wfKor,
            "r06": r06
        }
        return weather_info
    else:
        print(f"{category} 지역의 현재 시간에 해당하는 날씨 정보가 없습니다.")
        return None
    
def mk_xlsx(urls):
    current_path = "excel_files"
    weather_data = fetch_weather_data(urls)
    if weather_data:
        category = weather_data['category']
        file_name = f"{now}_{category}날씨.xlsx"
        file_path = os.path.join(current_path, file_name)

        wb = op.Workbook()
        sheet = wb.active
        sheet.title = "날씨 정보"

        
        sheet['A1'] = f"{category} 날씨 정보 ({now2})"
        sheet.merge_cells('A1:G1')
        sheet.merge_cells('A6:B6')
        sheet['A1'].font = Font(size=16)
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # 날씨 정보 데이터 입력해기
        sheet['A2'] = "기온"
        sheet['B2'] = f"{weather_data['temp']}°C"
        sheet['A3'] = "습도"
        sheet['B3'] = f"{weather_data['reh']}%"
        sheet['A4'] = "강수확률"
        sheet['B4'] = f"{weather_data['pop']}%"
        sheet['A5'] = "기상 상태"
        sheet['B5'] = f"{weather_data['wfKor']}"
        sheet['A6'] = "6시간 예상 강수량"
        sheet['C6'] = f"{weather_data['r06']}mm"
        sheet['A7'] = f"{now} 기사 제목 요약"
        sheet.merge_cells('A7:G7')
        sheet['A7'].font = Font(size=16)
        sheet['A7'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['H7'].hyperlink = f"{(url2)}"
        sheet['H7'].font = Font(color="0000FF", underline="single")
        
        
        if float(weather_data['r06']) >= 10:
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            sheet['C6'].fill = red_fill
            
        start_row = 8
        for idx, content in enumerate(filtered_a, start=start_row):
            sheet[f'A{idx}'] = content

       
        wb.save(file_path)
        print(f"{category} 지역의 날씨 정보가 저장되었습니다: {file_path}")
        print("\n")



